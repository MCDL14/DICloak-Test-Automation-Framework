from __future__ import annotations

import time
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def full_path(directory: str | Path, file_name: str | Path) -> Path:
    return Path(directory) / Path(file_name)


def batch_import_file(config: dict[str, Any]) -> Path:
    node = config["test_data"]["batch_import"]
    return full_path(node.get("file_dir", ""), node.get("file_name", ""))


def batch_export_file(config: dict[str, Any]) -> Path:
    node = config["test_data"]["batch_export"]
    return full_path(node.get("export_dir", ""), node.get("export_file_name", ""))


def member_export_file(config: dict[str, Any]) -> Path:
    node = config["test_data"]["member_export"]
    file_name = str(node.get("export_file_name", "")).strip()
    if file_name:
        return full_path(node.get("export_dir", ""), file_name)
    return latest_member_export_file(config)


def member_export_file_regex(config: dict[str, Any]) -> re.Pattern[str]:
    node = config["test_data"]["member_export"]
    pattern = str(node.get("export_file_regex") or r"^导出成员列表 - \d{12}\.xlsx$")
    return re.compile(pattern)


def latest_member_export_file(config: dict[str, Any]) -> Path:
    node = config["test_data"]["member_export"]
    export_dir = Path(node.get("export_dir", ""))
    regex = member_export_file_regex(config)
    matched = find_files_by_regex(export_dir, regex)
    if not matched:
        raise FileNotFoundError(f"no member export file matched in {export_dir}: {regex.pattern}")
    return max(matched, key=lambda item: item.stat().st_mtime)


def local_extension_file(config: dict[str, Any]) -> Path:
    node = config["test_data"]["local_extension"]
    package_path = Path(node.get("package_path", ""))
    package_name = str(node.get("package_name", "")).strip()
    if package_path.is_dir() or package_path.suffix == "":
        return package_path / package_name
    return package_path


def assert_existing_path(path: str | Path) -> Path:
    resolved = Path(path)
    if not resolved.exists():
        raise FileNotFoundError(f"path does not exist: {resolved}")
    return resolved


def wait_for_file(path: str | Path, timeout_seconds: int = 30, stable_seconds: int = 2) -> Path:
    file_path = Path(path)
    deadline = time.time() + timeout_seconds
    last_size = None
    stable_since = None

    while time.time() < deadline:
        if file_path.is_file():
            size = file_path.stat().st_size
            if size == last_size:
                stable_since = stable_since or time.time()
                if time.time() - stable_since >= stable_seconds:
                    return file_path
            else:
                last_size = size
                stable_since = None
        time.sleep(0.5)
    raise TimeoutError(f"file was not generated or did not become stable: {file_path}")


def find_files_by_regex(directory: str | Path, regex: re.Pattern[str]) -> list[Path]:
    dir_path = Path(directory)
    if not dir_path.is_dir():
        return []
    return [path for path in dir_path.iterdir() if path.is_file() and regex.fullmatch(path.name)]


def wait_for_file_matching(
    directory: str | Path,
    pattern: str,
    timeout_seconds: int = 30,
    stable_seconds: int = 2,
) -> Path:
    regex = re.compile(pattern)
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        matched = find_files_by_regex(directory, regex)
        if matched:
            latest = max(matched, key=lambda item: item.stat().st_mtime)
            return wait_for_file(latest, timeout_seconds=max(1, int(deadline - time.time())), stable_seconds=stable_seconds)
        time.sleep(0.5)
    raise TimeoutError(f"no file matched in {directory}: {pattern}")


def read_xlsx_headers(path: str | Path, sheet_name: str | None = None) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(cell).strip() if cell is not None else "" for cell in row]
    finally:
        workbook.close()


def read_xlsx_rows(path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data: list[dict[str, Any]] = []
        for row in rows[1:]:
            data.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
        return data
    finally:
        workbook.close()


def write_xlsx_rows(path: str | Path, rows: list[dict[str, Any]]) -> Path:
    file_path = Path(path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    if not rows:
        workbook.save(file_path)
        workbook.close()
        return file_path

    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    workbook.save(file_path)
    workbook.close()
    return file_path


def generate_bookmark_files(config: dict[str, Any]) -> tuple[Path, Path]:
    bookmark = config["test_data"]["bookmark"]
    storage_dir = Path(bookmark.get("storage_dir", ""))
    overwrite_path = storage_dir / bookmark.get("overwrite_file_name", "")
    append_path = storage_dir / bookmark.get("append_file_name", "")
    write_tabular_file(overwrite_path, list(bookmark.get("overwrite_rows", [])))
    write_tabular_file(append_path, list(bookmark.get("append_rows", [])))
    return overwrite_path, append_path


def write_tabular_file(path: str | Path, rows: list[Any]) -> Path:
    file_path = Path(path)
    if file_path.suffix.lower() == ".xlsx":
        normalized_rows = [_normalize_row(row) for row in rows]
        return write_xlsx_rows(file_path, normalized_rows)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_text("\n".join(_row_to_text(row) for row in rows), encoding="utf-8")
    return file_path


def _normalize_row(row: Any) -> dict[str, Any]:
    if isinstance(row, dict):
        return row
    return {"value": row}


def _row_to_text(row: Any) -> str:
    if isinstance(row, dict):
        return "\t".join(str(value) for value in row.values())
    return str(row)


class CreatedDataRegistry:
    def __init__(self) -> None:
        self.records: list[tuple[str, str]] = []

    def add(self, data_type: str, identifier: str) -> None:
        self.records.append((data_type, identifier))

    def clear(self) -> None:
        self.records.clear()

    def snapshot(self) -> list[tuple[str, str]]:
        return list(self.records)


class CleanupManager:
    def __init__(self) -> None:
        self.registry = CreatedDataRegistry()
        self.handlers: dict[str, Any] = {}
        self.failures: list[str] = []

    def register_handler(self, data_type: str, handler) -> None:
        self.handlers[data_type] = handler

    def add_created(self, data_type: str, identifier: str) -> None:
        self.registry.add(data_type, identifier)

    def cleanup(self) -> bool:
        ok = True
        for data_type, identifier in reversed(self.registry.snapshot()):
            handler = self.handlers.get(data_type)
            if not handler:
                self.failures.append(f"no cleanup handler for {data_type}: {identifier}")
                ok = False
                continue
            try:
                handler(identifier)
            except Exception as exc:
                self.failures.append(f"cleanup failed for {data_type}:{identifier}: {exc}")
                ok = False
        if ok:
            self.registry.clear()
        return ok
