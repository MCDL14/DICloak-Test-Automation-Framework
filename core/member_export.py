from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MEMBER_EXPORT_HEADERS = (
    "成员ID",
    "成员名称",
    "登录账号/邮箱",
    "备注",
    "授权环境分组",
    "成员身份",
    "所属成员分组",
    "状态",
    "开启到期停用",
    "过期时间",
    "时区",
    "最近登录时间",
    "创建人",
    "创建时间",
)

AUTHORITY_LABELS = {
    "MANAGER": "经理",
    "ADMIN": "管理员",
    "MEMBER": "员工",
}

STATUS_LABELS = {
    "ENABLED": "启用中",
    "DISABLED": "已停用",
}


def extract_target_member_records(
    response_body: str,
    target_names: list[str] | tuple[str, ...],
) -> list[dict[str, Any]]:
    try:
        payload = json.loads(response_body or "{}")
    except json.JSONDecodeError as exc:
        raise AssertionError(f"member list response is not valid JSON: {exc}") from exc
    if not isinstance(payload, dict):
        raise AssertionError("member list response payload must be an object")
    if payload.get("code") != 0:
        raise AssertionError(
            f"member list response code mismatch: code={payload.get('code')}, msg={payload.get('msg')}"
        )

    data = payload.get("data")
    rows = data.get("list") if isinstance(data, dict) else None
    if not isinstance(rows, list):
        raise AssertionError("member list response data.list must be an array")

    selected: list[dict[str, Any]] = []
    for target_name in target_names:
        clean_name = str(target_name or "").strip()
        matches = [
            row
            for row in rows
            if isinstance(row, dict) and str(row.get("name") or "").strip() == clean_name
        ]
        if len(matches) != 1:
            raise AssertionError(
                f"member list response must contain exactly one target member: "
                f"name={clean_name}, matched={len(matches)}"
            )
        member = dict(matches[0])
        if not str(member.get("id") or "").strip():
            raise AssertionError(f"member list response target has no id: name={clean_name}")
        selected.append(member)
    return selected


def expected_export_row_from_member(member: dict[str, Any]) -> dict[str, str]:
    env_group_value = _expected_environment_groups(member)
    authority = str(member.get("authority") or "").strip().upper()
    status = str(member.get("status") or "").strip().upper()
    return {
        "成员ID": _cell_text(member.get("id")),
        "成员名称": _cell_text(member.get("name")),
        "登录账号/邮箱": _cell_text(member.get("email")),
        "备注": _cell_text(member.get("remark")),
        "授权环境分组": env_group_value,
        "成员身份": AUTHORITY_LABELS.get(authority, authority),
        "所属成员分组": _cell_text(member.get("role_name")),
        "状态": STATUS_LABELS.get(status, status),
        "开启到期停用": "已开启" if _as_bool(member.get("disuse_enable")) else "已关闭",
        "过期时间": _cell_text(member.get("disuse_time")),
        "时区": _cell_text(member.get("time_zone")),
        "最近登录时间": _cell_text(member.get("last_login_time")) or "尚未登录",
        "创建人": _cell_text(member.get("create_by_name")),
        "创建时间": _cell_text(member.get("create_time")),
    }


def read_member_export_rows(file_path: Path | str) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(Path(file_path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], []
        headers = [_cell_text(value) for value in values[0]]
        rows: list[dict[str, str]] = []
        for values_row in values[1:]:
            row = {
                header: _cell_text(values_row[index] if index < len(values_row) else None)
                for index, header in enumerate(headers)
                if header
            }
            if any(row.values()):
                rows.append(row)
        return headers, rows
    finally:
        workbook.close()


def assert_member_export_matches_api(
    headers: list[str],
    actual_rows: list[dict[str, str]],
    member_records: list[dict[str, Any]],
) -> None:
    expected_headers = list(MEMBER_EXPORT_HEADERS)
    if headers != expected_headers:
        raise AssertionError(
            f"member export headers mismatch: actual={headers}, expected={expected_headers}"
        )
    if len(actual_rows) != len(member_records):
        raise AssertionError(
            f"member export row count mismatch: actual={len(actual_rows)}, expected={len(member_records)}"
        )

    expected_rows = [expected_export_row_from_member(member) for member in member_records]
    actual_by_name = _rows_by_unique_name(actual_rows, "export")
    expected_by_name = _rows_by_unique_name(expected_rows, "member API")
    if set(actual_by_name) != set(expected_by_name):
        raise AssertionError(
            "member export scope mismatch: "
            f"actual={sorted(actual_by_name)}, expected={sorted(expected_by_name)}"
        )

    for member_name, expected_row in expected_by_name.items():
        actual_row = actual_by_name[member_name]
        for header in MEMBER_EXPORT_HEADERS:
            actual_value = _cell_text(actual_row.get(header))
            expected_value = _cell_text(expected_row.get(header))
            if header == "授权环境分组" and expected_value != "全部分组":
                matched = _normalized_environment_groups(actual_value) == _normalized_environment_groups(
                    expected_value
                )
            else:
                matched = actual_value == expected_value
            if not matched:
                raise AssertionError(
                    f"member export field mismatch: member={member_name}, field={header}, "
                    f"actual={actual_value!r}, expected={expected_value!r}"
                )


def _expected_environment_groups(member: dict[str, Any]) -> str:
    if _as_bool(member.get("all_env_group")):
        return "全部分组"
    env_groups = member.get("env_group_list")
    if not isinstance(env_groups, list):
        return ""
    names = [
        str(item.get("env_group_name") or "").strip()
        for item in env_groups
        if isinstance(item, dict) and str(item.get("env_group_name") or "").strip()
    ]
    return "、".join(names)


def _normalized_environment_groups(value: str) -> tuple[str, ...]:
    return tuple(
        sorted(
            {
                item.strip()
                for item in re.split(r"[\r\n,，、;；]+", str(value or ""))
                if item.strip()
            }
        )
    )


def _rows_by_unique_name(
    rows: list[dict[str, str]],
    source: str,
) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        member_name = _cell_text(row.get("成员名称"))
        if not member_name:
            raise AssertionError(f"{source} member row has empty 成员名称: {row}")
        if member_name in result:
            raise AssertionError(f"{source} contains duplicate member name: {member_name}")
        result[member_name] = row
    return result


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()
