from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

from core.assertions import assert_equal, assert_file_exists, assert_true
from core.cdp_driver import CDPDriver
from core.config import load_config, timeout_seconds
from core.files import batch_export_file, wait_for_file
from core.logger import setup_logger
from core.ui_driver import UIDriver
from pages.environment_page import EnvironmentPage
from pages.login_page import LoginPage


CASE_MODULE = "环境管理"


class TestExportEnvironment(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.config = load_config(Path("config/config.yaml"))
        cls.logger = setup_logger(cls.config)
        cls.cdp = CDPDriver(cls.config, cls.logger)
        cls.cdp.connect()
        LoginPage(cdp_driver=cls.cdp, config=cls.config).ensure_logged_in_as_config_account()

    @classmethod
    def tearDownClass(cls) -> None:
        cls.cdp.close()

    def test_export_selected_environments(self) -> None:
        keyword = "导出环境"
        export_file = batch_export_file(self.config)
        export_timeout = timeout_seconds(self.config, "batch_export_seconds", 120)
        download_event_timeout = min(export_timeout, 30)
        environment_page = EnvironmentPage(cdp_driver=self.cdp, config=self.config)

        try:
            export_file.parent.mkdir(parents=True, exist_ok=True)
            if export_file.exists():
                export_file.unlink()

            environment_page.open_list()
            environment_page.search_environment(keyword)
            expected_infos = environment_page.environment_infos_in_current_list()
            assert_true(expected_infos, f"no environments were found for export: keyword={keyword}")
            selected_count = environment_page.select_all_environments_in_current_list()
            assert_true(selected_count > 0, f"no environments were selected for export: keyword={keyword}")
            assert_equal(
                selected_count,
                len(expected_infos),
                f"selected environment count did not match current list: selected={selected_count}, rows={expected_infos}",
            )

            environment_page.open_export_selected_environments_dialog()
            try:
                environment_page.confirm_export_environment_and_save_download(
                    export_file,
                    timeout_seconds=download_event_timeout,
                )
            except PlaywrightTimeoutError:
                ui_driver = UIDriver(self.config, self.logger)
                ui_driver.save_file_in_dialog(export_file, timeout=15)

            generated_file = wait_for_file(export_file, timeout_seconds=export_timeout)
            assert_file_exists(generated_file, f"export file was not generated: {generated_file}")
            assert_true(generated_file.stat().st_size > 0, f"export file is empty: {generated_file}")
            exported_infos = self._exported_environment_infos(generated_file)
            assert_equal(
                self._sort_environment_infos(exported_infos),
                self._sort_environment_infos(expected_infos),
                f"exported xlsx content did not match selected environments: file={generated_file}",
            )
        finally:
            try:
                environment_page.clear_search()
            except Exception:
                pass

    def _exported_environment_infos(self, file_path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            assert_true(len(rows) >= 4, f"exported xlsx has no data rows: {file_path}")
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            required_headers = ["环境序号", "环境名称", "环境分组", "备注"]
            for header in required_headers:
                assert_true(header in headers, f"exported xlsx missing header {header}: headers={headers}")

            header_index = {header: headers.index(header) for header in required_headers}
            infos: list[dict[str, str]] = []
            for row in rows[3:]:
                serial = self._cell_text(row, header_index["环境序号"])
                name = self._cell_text(row, header_index["环境名称"])
                if not serial and not name:
                    continue
                infos.append(
                    {
                        "serial": self._normalize_serial(serial),
                        "name": name,
                        "group": self._cell_text(row, header_index["环境分组"]),
                        "remark": self._cell_text(row, header_index["备注"]),
                    }
                )
            return infos
        finally:
            workbook.close()

    def _sort_environment_infos(self, infos: list[dict[str, str]]) -> list[dict[str, str]]:
        return sorted(
            [
                {
                    "serial": self._normalize_serial(item.get("serial", "")),
                    "name": str(item.get("name", "")).strip(),
                    "group": str(item.get("group", "")).strip(),
                    "remark": str(item.get("remark", "")).strip(),
                }
                for item in infos
            ],
            key=lambda item: (item["serial"], item["name"]),
        )

    def _cell_text(self, row: tuple, index: int) -> str:
        if index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        return str(value).strip()

    def _normalize_serial(self, value: str) -> str:
        text = str(value).strip()
        if text.endswith(".0"):
            text = text[:-2]
        return text


if __name__ == "__main__":
    unittest.main()
