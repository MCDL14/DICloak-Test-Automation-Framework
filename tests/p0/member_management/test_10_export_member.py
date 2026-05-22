from __future__ import annotations

import re
import unittest
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

from core.assertions import assert_file_exists, assert_true
from core.cdp_driver import CDPDriver
from core.config import load_config, timeout_seconds
from core.files import wait_for_file
from core.logger import setup_logger
from pages.login_page import LoginPage
from pages.member_page import MemberPage


CASE_MODULE = "成员管理"


class TestExportMember(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.config = load_config(Path("config/config.yaml"))
        cls.logger = setup_logger(cls.config)
        cls.cdp = CDPDriver(cls.config, cls.logger)
        cls.cdp.connect()
        LoginPage(cdp_driver=cls.cdp, config=cls.config).ensure_logged_in_as_config_account()

    @classmethod
    def tearDownClass(cls) -> None:
        cls.cdp.close()

    def test_export_selected_members(self) -> None:
        target_member_names = ["自动化成员1", "外部成员1"]
        member_page = MemberPage(cdp_driver=self.cdp, config=self.config)

        # 读取测试数据配置
        member_export_cfg = self.config["test_data"]["member_export"]
        expected_file = Path(member_export_cfg["expected_file_full_path"])
        export_dir = Path(member_export_cfg["export_dir"])

        export_timeout = timeout_seconds(self.config, "batch_export_seconds", 120)
        download_event_timeout = min(export_timeout, 30)

        # 导出文件统一保存到临时路径，避免与正则匹配双轨
        exported_file: Path | None = None
        temp_download_path = export_dir / "_member_export_tmp.xlsx"

        try:
            # Step 1: 选中"自动化成员1"和"外部成员1"
            member_page.open_list()
            member_page.clear_filters()

            # 先逐个按成员名称精确筛选获取目标 ID，避免依赖当前页、默认分页或排序。
            target_member_ids: list[str] = []
            for name in target_member_names:
                target_member_ids.append(member_page.member_id_by_exact_name(name))
                member_page.clear_filters()
            member_page.select_members_by_ids(target_member_ids)

            # Step 2-4: 点击"更多操作" -> "导出成员" -> "导出所选"，并捕获下载
            # 成员导出点击"导出所选"后直接触发下载，不存在 .el-dialog 确认弹窗
            export_dir.mkdir(parents=True, exist_ok=True)
            if temp_download_path.exists():
                temp_download_path.unlink()

            suggested_filename = ""
            try:
                suggested_filename = member_page.export_selected_members_and_save_download(
                    temp_download_path,
                    timeout_seconds=download_event_timeout,
                )
            except PlaywrightTimeoutError:
                # CDP 下载未捕获到时，回退到 Windows 系统保存弹窗
                # UIDriver 会按 temp_download_path 写入文件名
                member_page.export_selected_members_via_save_dialog(temp_download_path)

            if suggested_filename:
                filename_regex = str(member_export_cfg.get("export_file_regex") or "").strip()
                if filename_regex:
                    assert_true(
                        re.fullmatch(filename_regex, suggested_filename) is not None,
                        f"export filename mismatch: actual={suggested_filename}, regex={filename_regex}",
                    )

            generated_file = wait_for_file(temp_download_path, timeout_seconds=export_timeout)
            assert_file_exists(generated_file, f"export file was not generated: {generated_file}")
            assert_true(generated_file.stat().st_size > 0, f"export file is empty: {generated_file}")
            exported_file = generated_file

            # Step 5: 比较导出文件和预置的正确导出文件
            self._assert_xlsx_content_match(exported_file, expected_file, target_member_names)

        finally:
            # Step 6: 清理已勾选成员、删除导出文件、恢复列表筛选状态
            try:
                member_page.clear_selected_members()
            except Exception:
                pass
            if exported_file and exported_file.exists():
                try:
                    exported_file.unlink()
                except Exception:
                    pass
            if temp_download_path.exists():
                try:
                    temp_download_path.unlink()
                except Exception:
                    pass
            try:
                member_page.open_list()
                member_page.clear_filters()
            except Exception:
                pass

    def _assert_xlsx_content_match(
        self,
        actual_file: Path,
        expected_file: Path,
        member_names: list[str],
    ) -> None:
        """比较导出文件与预置文件：表头一致、选中成员行均存在且内容一致。"""
        actual_rows = self._read_xlsx_rows(actual_file)
        expected_rows = self._read_xlsx_rows(expected_file)

        assert_true(bool(actual_rows), f"exported file has no data rows: {actual_file}")
        assert_true(bool(expected_rows), f"expected file has no data rows: {expected_file}")

        # 表头一致
        actual_headers = list(actual_rows[0].keys())
        expected_headers = list(expected_rows[0].keys())
        assert_true(
            actual_headers == expected_headers,
            f"xlsx headers mismatch: actual={actual_headers}, expected={expected_headers}",
        )

        # 选中成员在导出文件中均存在；用显式列名 "成员名称" 而不是依赖列顺序
        name_header = "成员名称" if "成员名称" in actual_headers else actual_headers[1]
        actual_names = {
            str(row.get(name_header, "")).strip()
            for row in actual_rows
            if str(row.get(name_header, "")).strip()
        }
        expected_names = set(member_names)
        assert_true(
            actual_names == expected_names,
            f"exported member scope mismatch: actual={sorted(actual_names)}, expected={sorted(expected_names)}",
        )
        for name in member_names:
            found = any(
                str(row.get(name_header, "")).strip() == name
                for row in actual_rows
            )
            assert_true(
                found,
                f"member '{name}' was not found in exported file: {actual_file}",
            )

        # 动态字段（导出时间、登录时间等会随时间变化），需要与实际导出表头保持一致
        dynamic_skip_headers = {"成员ID", "最近登录时间", "创建时间"}

        # 选中成员行内容与预置文件一致
        for name in member_names:
            actual_row = self._find_row_by_member_name(actual_rows, name, name_header)
            expected_row = self._find_row_by_member_name(expected_rows, name, name_header)
            assert_true(
                bool(actual_row),
                f"actual row for member '{name}' not found",
            )
            assert_true(
                bool(expected_row),
                f"expected row for member '{name}' not found in expected file",
            )
            for header in expected_headers:
                if header in dynamic_skip_headers:
                    continue
                actual_value = str(actual_row.get(header, "")).strip()
                expected_value = str(expected_row.get(header, "")).strip()
                assert_true(
                    actual_value == expected_value,
                    f"member '{name}' column '{header}' mismatch: actual='{actual_value}', expected='{expected_value}'",
                )

    def _read_xlsx_rows(self, file_path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            data: list[dict[str, str]] = []
            for row in rows[1:]:
                item = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        item[headers[idx]] = str(value).strip() if value is not None else ""
                # 跳过空行
                if any(v for v in item.values()):
                    data.append(item)
            return data
        finally:
            workbook.close()

    def _find_row_by_member_name(
        self,
        rows: list[dict[str, str]],
        member_name: str,
        name_header: str,
    ) -> dict[str, str] | None:
        for row in rows:
            if str(row.get(name_header, "")).strip() == member_name:
                return row
        return None


if __name__ == "__main__":
    unittest.main()
